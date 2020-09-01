from flask import Flask, render_template, request, redirect, url_for, session, make_response
from flask_mail import Mail, Message
import sqlite3
import os
from flask_bootstrap import Bootstrap
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, BooleanField
from wtforms.validators import InputRequired, Length, Email
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import datetime
import time
from time import strftime
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Color, colors, PatternFill, Border
from openpyxl.styles.borders import Border, Side
import numpy as np
from numpy import array
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.workbook.protection import WorkbookProtection
from copy import copy


app = Flask(__name__)
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'
app.config['SQLALCHEMY_DATABASE_URI'] = \
    'sqlite:///diacompanion.db'
app.config['TESTING'] = False
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = 'teos.sicrets@gmail.com'
app.config['MAIL_PASSWORD'] = 'ShurfLL1'
app.config['MAIL_DEFAULT_SENDER'] = ('Еженедельник', 'teos.sicrets@gmail.com')
app.config['MAIL_MAX_EMAILS'] = None
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAIL_ASCII_ATTACHMENTS'] = False
app.config['SESSION_COOKIE_SAMESITE'] = "Lax"

Bootstrap(app)
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
mail = Mail(app)


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(15), unique=True)
    username1 = db.Column(db.String(15), unique=True)
    email = db.Column(db.String(80), unique=True)
    password = db.Column(db.String(15))


class LoginForm(FlaskForm):
    username = StringField('Имя пользователя',
                           validators=[InputRequired(), Length(min=5, max=15)])
    password = PasswordField('Пароль', validators=[InputRequired(),
                                                   Length(min=8, max=15)])
    remember = BooleanField('Запомнить меня')


class RegisterForm(FlaskForm):
    email = StringField('Email', validators=[InputRequired(),
                                             Email(message='Invalid email'),
                                             Length(max=50)])
    username = StringField('Никнейм пользователя', validators=[InputRequired(),
                                                           Length(min=5,
                                                                  max=15)])
    username1 = StringField('ФИО пользователя', validators=[InputRequired(),
                                                           Length(min=5,
                                                                  max=15)])                                                                  
    password = PasswordField('Пароль', validators=[InputRequired(),
                                                   Length(min=8, max=15)])


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route('/')
def zero():
    # Перенаправляем на страницу входа/регистрации
    return redirect(url_for('login'))


@app.route('/news')
@login_required
def news():
    # Главная страница
    session['username'] = current_user.username
    session['user_id'] = current_user.id
    session['date'] = datetime.datetime.today().date()
    path = os.path.dirname(os.path.abspath(__file__))
    db = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db)
    cur = con.cursor()
    cur.execute("""SELECT food,libra FROM basket WHERE user_id = ?""",(session['user_id'],))
    result = cur.fetchall()

    return render_template("searching.html", result=result)


@app.route('/search_page')
@login_required
def search_page():
    # Поисковая страница
    return render_template("searching.html")


@app.route('/searchlink/<string:search_string>')
@login_required
def searchlink(search_string):
    # Работа селекторного меню "выбрать категорию"
    path = os.path.dirname(os.path.abspath(__file__))
    db = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db)
    cur = con.cursor()
    cur.execute("""SELECT DISTINCT (name) name,_id FROM
                constant_food WHERE category LIKE ?
                GROUP BY name""", ('%{}%'.format(search_string),))
    result = cur.fetchall()
    con.close()

    return render_template('searching_add.html', result=result)


@app.route('/search', methods=['POST'])
@login_required
def search():
    # Основная функция сайта - поиск по базе данных
    if request.method == 'POST':
        search_string = request.form['input_query']
        path = os.path.dirname(os.path.abspath(__file__))
        db = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db)
        cur = con.cursor()

        cur.execute(""" SELECT category FROM constant_foodGroups""")
        category_a = cur.fetchall()
        if (request.form['input_query'],) in category_a:
            cur.execute('''SELECT name,_id FROM constant_food
                        WHERE category LIKE ?
                        GROUP BY name''', ('%{}%'.format(search_string),))
            result = cur.fetchall()
        else:
            cur.execute('''SELECT name,_id FROM constant_food
                        WHERE name LIKE ?
                        GROUP BY name''', ('%{}%'.format(search_string),))
            result = cur.fetchall()
        con.close()
        return render_template('searching_add.html', result=result)


@app.route('/login', methods=['GET', 'POST'])
def login():
    # Авторизация пользователя
    form = LoginForm()

    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user:
            if check_password_hash(user.password, form.password.data):
                login_user(user, remember=form.remember.data)
                return redirect(url_for('news'))
        return redirect(url_for('login'))

    return render_template('login.html', form=form)


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    # Регистрация пользователя
    form = RegisterForm()

    if form.validate_on_submit():
        hashed_password = generate_password_hash(form.password.data,
                                                 method='sha256')
        new_user = User(username=form.username.data, username1=form.username1.data, email=form.email.data,
                        password=hashed_password)
        db.session.add(new_user)
        db.session.commit()
        return redirect(url_for('login'))

    return render_template('signup.html', form=form)


@app.route('/logout')
@login_required
def logout():
    # Выход из сети
    path = os.path.dirname(os.path.abspath(__file__))
    db = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db)
    cur = con.cursor()
    cur.execute("""DELETE FROM basket WHERE user_id = ?""",(session['user_id'],))
    con.commit()
    con.close()
    logout_user()
    return redirect(url_for('login'))


@app.route('/favourites', methods=['POST', 'GET'])
@login_required
def favour():
    # Добавляем блюда в список
    if request.method == 'POST':

        L1 = request.form.getlist('row')
        libra = request.form['libra']

        path = os.path.dirname(os.path.abspath(__file__))
        db = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db)
        cur = con.cursor()

        for i in range(len(L1)):
            cur.execute("""INSERT INTO basket
                           VALUES(?,?,?)""", (session['user_id'], L1[i], libra))
            con.commit()
        con.close()

    return redirect(url_for('news'))


@app.route('/favourites_dell', methods=['POST','GET'])
@login_required
def favour_dell():
    if request.method == 'POST':
        flist = request.form.getlist('row')
        food = []
        libra = []
        for i in range(len(flist)):
            flist[i] = flist[i].split('//')
            food.append(flist[i][0])
            libra.append(flist[i][1])
  
        for i in range(len(food)):
            path = os.path.dirname(os.path.abspath(__file__))
            db = os.path.join(path, 'diacompanion.db')
            con = sqlite3.connect(db)
            cur = con.cursor()
            cur.execute("""DELETE FROM basket WHERE user_id = ? AND food = ? AND libra = ?""",(session['user_id'], food[i], libra[i]))
            con.commit()
            con.close()
    return redirect(url_for('news'))


@app.route('/favourites_add', methods=['POST', 'GET'])
@login_required
def favour_add():
    # Добавляем блюда в основную базу данных и стираем временный список basket 
    if request.method == 'POST':

        brf1 = datetime.time(7, 0)
        brf2 = datetime.time(11, 30)
        obed1 = datetime.time(12, 0)
        obed2 = datetime.time(15, 0)
        ujin1 = datetime.time(18, 0)
        ujin2 = datetime.time(22, 0)
        now = datetime.datetime.now().time()

        time = request.form['timer']
        if time == "":
            x = datetime.datetime.now().time()
            time = x.strftime("%R")
        else:
            x = datetime.datetime.strptime(time, "%H:%M")
            time = x.strftime("%R")

        date = request.form['calendar']
        if date == "":
            y = datetime.datetime.today().date()
            date = y.strftime("%d.%m.%Y")
            week_day = y.strftime("%A")
        else:
            y = datetime.datetime.strptime(date, "%Y-%m-%d")
            y = y.date()
            date = y.strftime("%d.%m.%Y")
            week_day = y.strftime("%A")

        if week_day == 'Monday':
            week_day = 'Понедельник'
        elif week_day == 'Tuesday':
            week_day = 'Вторник'
        elif week_day == 'Wednesday':
            week_day = 'Среда'
        elif week_day == 'Thursday':
            week_day = 'Четверг'
        elif week_day == 'Friday':
            week_day = 'Пятница'
        elif week_day == 'Saturday':
            week_day = 'Суббота'
        else:
            week_day = 'Воскресенье'
        typ = request.form['food_type']
        if typ == "Авто":
            if now < brf1 and now > brf2:
                typ = "Завтрак"
            elif now < obed2 and now > obed1:
                typ = "Обед"
            elif now < ujin2 and now > ujin1:
                typ = "Ужин"
            else:
                typ = "Перекус"

        path = os.path.dirname(os.path.abspath(__file__))
        db = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db)
        cur = con.cursor()

       # Достаем названия и граммы из временно созданной корзины basket 
        cur.execute("""SELECT food FROM basket WHERE user_id = ?""",(session['user_id'],))
        L1 = cur.fetchall()

        cur.execute("""SELECT libra FROM basket WHERE user_id = ?""",(session['user_id'],))
        libra = cur.fetchall()

       # Достаем все необходимые для диеты параметры
        for i in range(len(L1)):
            cur.execute('''SELECT prot FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            prot = cur.fetchall()
            cur.execute('''SELECT carbo FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            carbo = cur.fetchall()
            cur.execute('''SELECT fat FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            fat = cur.fetchall()
            cur.execute('''SELECT ec FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            ec = cur.fetchall()
            cur.execute('''SELECT water FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            water = cur.fetchall()
            cur.execute('''SELECT mds FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            mds = cur.fetchall()
            cur.execute('''SELECT kr FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            kr = cur.fetchall()
            cur.execute('''SELECT pv FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            pv = cur.fetchall()
            cur.execute('''SELECT ok FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            ok = cur.fetchall()
            cur.execute('''SELECT zola FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            zola = cur.fetchall()
            cur.execute('''SELECT na FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            na = cur.fetchall()
            cur.execute('''SELECT k FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            k = cur.fetchall()
            cur.execute('''SELECT ca FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            ca = cur.fetchall()
            cur.execute('''SELECT mg FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            mg = cur.fetchall()
            cur.execute('''SELECT p FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            p = cur.fetchall()
            cur.execute('''SELECT fe FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            fe = cur.fetchall()
            cur.execute('''SELECT a FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            a = cur.fetchall()
            cur.execute('''SELECT kar FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            kar = cur.fetchall()
            cur.execute('''SELECT re FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            re = cur.fetchall()
            cur.execute('''SELECT b1 FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            b1 = cur.fetchall()
            cur.execute('''SELECT b2 FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            b2 = cur.fetchall()
            cur.execute('''SELECT rr FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            rr = cur.fetchall()
            cur.execute('''SELECT c FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            c = cur.fetchall()
            cur.execute('''SELECT hol FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            hol = cur.fetchall()
            cur.execute('''SELECT nzhk FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            nzhk = cur.fetchall()
            cur.execute('''SELECT ne FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            ne = cur.fetchall()
            cur.execute('''SELECT te FROM constant_food
                        WHERE name = ?''', (L1[i][0],))
            te = cur.fetchall()
            for pr in prot:
                print(pr[0])
            for car in carbo:
                print(car[0])
            for fa in fat:
                print(fa[0])
            for energy in ec:
                print(energy[0])
            for wat in water:
                print(wat[0])
            for md in mds:
                print(md[0])
            for kr1 in kr:
                print(kr1[0])
            for pv1 in pv:
                print(pv1[0])
            for ok1 in ok:
                print(ok1[0])
            for zola1 in zola:
                print(zola1[0])
            for na1 in na:
                print(na1[0])
            for k1 in k:
                print(k1[0])
            for ca1 in ca:
                print(ca1[0])
            for mg1 in mg:
                print(mg1[0])
            for p1 in p:
                print(p1[0])
            for fe1 in fe:
                print(fe1[0])
            for a1 in a:
                print(a1[0])
            for kar1 in kar:
                print(kar1[0])
            for re1 in re:
                print(re1[0])
            for b11 in b1:
                print(b11[0])
            for b21 in b2:
                print(b21[0])
            for rr1 in rr:
                print(rr1[0])
            for c1 in c:
                print(c1[0])
            for hol1 in hol:
                print(hol1[0])
            for nzhk1 in nzhk:
                print(nzhk1[0])
            for ne1 in ne:
                print(ne1[0])
            for te1 in te:
                print(te1[0])
            pustoe = ''
            cur.execute("""INSERT INTO favourites
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
                        ?,?,?,?,?,?,?,?,?,?,?,?,?)""", (session['user_id'],
                        week_day,
                        date, time, typ, L1[i][0], libra[i][0],
                        str(pr[0]), str(car[0]), str(fa[0]), str(energy[0]),
                        pustoe, str(wat[0]), str(md[0]), str(kr1[0]),
                        str(pv1[0]), str(ok1[0]), str(zola1[0]), str(na1[0]),
                        str(k1[0]), str(ca1[0]), str(mg1[0]), str(p1[0]),
                        str(fe1[0]), str(a1[0]), str(kar1[0]), str(re1[0]),
                        str(b11[0]), str(b21[0]), str(rr1[0]),
                        str(c1[0]), str(hol1[0]), str(nzhk1[0]),
                        str(ne1[0]), str(te1[0])))
            con.commit()
        
        cur.execute("""DELETE FROM basket WHERE user_id = ?""",(session['user_id'],))
        con.commit()
        con.close()
    return redirect(url_for('news'))


@app.route('/activity')
@login_required
def activity():
    # Страница физической активности
    path = os.path.dirname(os.path.abspath(__file__))
    db = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db)
    cur = con.cursor()
    cur.execute("""SELECT date,time,min,type,user_id
                    FROM activity WHERE user_id = ?""", (session['user_id'],))
    Act = cur.fetchall()
    cur.execute("""SELECT date,time,hour,type,user_id
                    FROM sleep WHERE user_id = ?""", (session['user_id'],))    
    Sleep = cur.fetchall()
    con.close()
    return render_template('activity.html', Act=Act, Sleep=Sleep)


@app.route('/add_activity', methods=['POST'])
@login_required
def add_activity():
    # Добавляем нагрузку в базу данных
    if request.method == 'POST':
        #td = datetime.datetime.today().date()
        #date = td.strftime("%d.%m.%Y")
        date = datetime.datetime.strptime(request.form['calendar'], "%Y-%m-%d")
        date = date.strftime("%d.%m.%Y")

        min1 = request.form['min']
        type1 = request.form['type1']
        if type1 == '1':
            type1 = 'Ходьба'
        elif type1 == '2':
            type1 = 'Зарядка'
        elif type1 == '3':
            type1 = 'Спорт'
        elif type1 == '4':
            type1 = 'Уборка в квартире'
        elif type1 == '5':
            type1 = 'Работа в огороде'
        else:
            type1 = 'Сон'
        time1 = request.form['timer']
        path = os.path.dirname(os.path.abspath(__file__))
        db = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db)
        cur = con.cursor()

        if type1 == 'Сон':    
            cur.execute("""INSERT INTO sleep (user_id,date,time,hour,type) VALUES(?,?,?,?,?)""",
                        (session['user_id'], date, time1, min1, type1))
        else:
            cur.execute("""INSERT INTO activity (user_id,date,time,min,type,empty) VALUES(?,?,?,?,?,?)""",
                        (session['user_id'], date, time1, min1, type1,' '))                            
        con.commit()
        con.close()

    return redirect(url_for('activity'))


@app.route('/lk')
@login_required
def lk():
    # Выводим названия блюд (дневник на текущую неделю)
    td = datetime.datetime.today().date()
    if td.strftime("%A") == 'Monday':
        delta = datetime.timedelta(0)
    elif td.strftime("%A") == 'Tuesday':
        delta = datetime.timedelta(1)
    elif td.strftime("%A") == 'Wednesday':
        delta = datetime.timedelta(2)
    elif td.strftime("%A") == 'Thursday':
        delta = datetime.timedelta(3)
    elif td.strftime("%A") == 'Friday':
        delta = datetime.timedelta(4)
    elif td.strftime("%A") == 'Saturday':
        delta = datetime.timedelta(5)
    else:
        delta = datetime.timedelta(6)

    m = td - delta
    M = m.strftime("%d.%m.%Y")
    t = m + datetime.timedelta(1)
    T = t.strftime("%d.%m.%Y")
    w = m + datetime.timedelta(2)
    W = w.strftime("%d.%m.%Y")
    tr = m + datetime.timedelta(3)
    TR = tr.strftime("%d.%m.%Y")
    fr = m + datetime.timedelta(4)
    FR = fr.strftime("%d.%m.%Y")
    st = m + datetime.timedelta(5)
    ST = st.strftime("%d.%m.%Y")
    sd = m + datetime.timedelta(6)
    SD = sd.strftime("%d.%m.%Y")

    path = os.path.dirname(os.path.abspath(__file__))
    db = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db)
    cur = con.cursor()

    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Понедельник',
                                  'Завтрак', M))
    MondayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""",
                (session['user_id'], 'Понедельник', 'Обед', M))
    MondayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""",
                (session['user_id'], 'Понедельник', 'Ужин', M))
    MondayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Понедельник',
                                  'Перекус', M))
    MondayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Вторник',
                                  'Завтрак', T))
    TuesdayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Вторник',
                                  'Обед', T))
    TuesdayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Вторник',
                                  'Ужин', T))
    TuesdayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Вторник',
                                  'Перекус', T))
    TuesdayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Среда',
                                  'Завтрак', W))
    WednesdayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Среда',
                                  'Обед', W))
    WednesdayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Среда',
                                  'Ужин', W))
    WednesdayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Среда',
                                  'Перекус', W))
    WednesdayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Четверг',
                                  'Завтрак', TR))
    ThursdayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Четверг',
                                  'Обед', TR))
    ThursdayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Четверг',
                                  'Ужин', TR))
    ThursdayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Четверг',
                                  'Перекус', TR))
    ThursdayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Пятница',
                                  'Завтрак', FR))
    FridayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Пятница',
                                  'Обед', FR))
    FridayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Пятница',
                                  'Ужин', FR))
    FridayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date= ?""", (session['user_id'], 'Пятница',
                                 'Перекус', FR))
    FridayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Суббота',
                                  'Завтрак', ST))
    SaturdayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Суббота',
                                  'Обед', ST))
    SaturdayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Суббота',
                                  'Ужин', ST))
    SaturdayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Суббота',
                                  'Перекус', ST))
    SaturdayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type =?
                AND date = ?""", (session['user_id'], 'Воскресенье',
                                  'Завтрак', SD))
    SundayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type =?
                AND date = ?""", (session['user_id'], 'Воскресенье',
                                  'Обед', SD))
    SundayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type =?
                AND date = ?""", (session['user_id'], 'Воскресенье',
                                  'Ужин', SD))
    SundayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type =?
                AND date = ?""", (session['user_id'], 'Воскресенье',
                                  'Перекус', SD))
    SundayP = cur.fetchall()
    con.close()

    return render_template('bootstrap_lk.html', name=session['username'],
                           MondayZ=MondayZ,
                           MondayO=MondayO,
                           MondayY=MondayY,
                           MondayP=MondayP,
                           TuesdayZ=TuesdayZ,
                           TuesdayO=TuesdayO,
                           TuesdayY=TuesdayY,
                           TuesdayP=TuesdayP,
                           WednesdayZ=WednesdayZ,
                           WednesdayO=WednesdayO,
                           WednesdayY=WednesdayY,
                           WednesdayP=WednesdayP,
                           ThursdayZ=ThursdayZ,
                           ThursdayO=ThursdayO,
                           ThursdayY=ThursdayY,
                           ThursdayP=ThursdayP,
                           FridayZ=FridayZ,
                           FridayO=FridayO,
                           FridayY=FridayY,
                           FridayP=FridayP,
                           SaturdayZ=SaturdayZ,
                           SaturdayO=SaturdayO,
                           SaturdayY=SaturdayY,
                           SaturdayP=SaturdayP,
                           SundayZ=SundayZ,
                           SundayO=SundayO,
                           SundayY=SundayY,
                           SundayP=SundayP,
                           m=M,
                           t=T,
                           w=W,
                           tr=TR,
                           fr=FR,
                           st=ST,
                           sd=SD)


@app.route('/delete', methods=['POST'])
@login_required
def delete():
    # Удаление данных из дневника приемов пищи за неделю
    if request.method == 'POST':
        path = os.path.dirname(os.path.abspath(__file__))
        db = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db)
        cur = con.cursor()
        L = request.form.getlist('checked')
        print(L)
        for i in range(len(L)):
            L1 = L[i].split('//')
            print(L1)
            cur.execute('''DELETE FROM favourites WHERE food = ?
                        AND date = ?
                        AND time = ?
                        AND type = ?
                        AND user_id = ?''', (L1[0], L1[1], L1[2], L1[3],
                                             session['user_id']))
        con.commit()
        con.close()
    return redirect(url_for('lk'))


@app.route('/remove', methods=['POST'])
@login_required
def remove():
    # Удаление данных из физической активности за неделю
    if request.method == 'POST':
        path = os.path.dirname(os.path.abspath(__file__))
        db = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db)
        cur = con.cursor()
        L = request.form.getlist('selected')
        for i in range(len(L)):
            L1 = L[i].split('/')
            if L1[3] != 'Сон':
                cur.execute('''DELETE FROM activity WHERE date = ?
                            AND time = ?
                            AND min = ?
                            AND type = ?
                            AND user_id = ?''', (L1[0], L1[1], L1[2], L1[3],
                                                 session['user_id'])) 
            else:
                cur.execute('''DELETE FROM sleep WHERE date = ?
                            AND time = ?
                            AND hour = ?
                            AND type = ?
                            AND user_id = ?''', (L1[0], L1[1], L1[2], L1[3],
                                                 session['user_id']))                 
        con.commit()
        con.close()
    return redirect(url_for('activity'))


@app.route('/arch')
@login_required
def arch():
    # Архив за все время
    path = os.path.dirname(os.path.abspath(__file__))
    db = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db)
    cur = con.cursor()
    cur.execute(
        """SELECT week_day,date,time,food,libra,type,prot,carbo,fat,energy
           FROM favourites WHERE user_id = ?""", (session['user_id'],))
    L = cur.fetchall()
    con.close()
    return render_template('arch.html', L=L)


@app.route('/email', methods=['GET', 'POST'])
@login_required
def email():
    # Отправляем отчет по почте отчет
    if request.method == 'POST':
        # Получили список имейлов на которые надо отправить
        mail1 = request.form.getlist('email_sendto')

        # Получили все необходимые данные из базы данных
        path = os.path.dirname(os.path.abspath(__file__))
        db = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db)
        cur = con.cursor()
        cur.execute('''SELECT date,time,type,
                    food,libra,carbo,prot,
                    fat,energy,micr,water,mds,kr,pv,ok,
                    zola,na,k,ca,mg,p,fe,a,kar,re,b1,b2,
                    rr,c,hol,nzhk,ne,te FROM favourites
                    WHERE user_id = ?''', (session['user_id'],))
        L = cur.fetchall()
        cur.execute('''SELECT date,time,min,type,empty FROM activity
                       WHERE user_id = ?''', (session['user_id'],))
        L1 = cur.fetchall()
        cur.execute('''SELECT date,time,hour FROM sleep
                        WHERE user_id =?''', (session['user_id'],))
        L2 = cur.fetchall()

        cur.execute('''SELECT DISTINCT date FROM
                        favourites WHERE user_id = ?''', (session['user_id'],))
        date = cur.fetchall()
        cur.execute('''SELECT username1 FROM user WHERE id = ?''', (session['user_id'],))
        fio = cur.fetchall()
        con.close()

        # Приемы пищи
        food_weight = pd.DataFrame(L, columns=['Дата', 'Время', 'Прием пищи',
                                               'Продукт', 'Масса, гр',
                                               'Углеводы, гр',
                                               'Белки, гр', 'Жиры, гр',
                                               'ККал',
                                               'Микроэлементы', 'Вода, в г', 'МДС, в г',
                                               'Крахмал, в г', 'Пищ вол, в г',
                                               'Орган кислота, в г', 'Зола, в г',
                                               'Натрий, в мг', 'Калий, в мг', 'Кальций, в мг',
                                               'Магний, в мг', 'Фосфор, в мг', 'Железо, в мг',
                                               'Ретинол, в мкг', 'Каротин, в мкг',
                                               'Ретин экв, в мкг', 'Тиамин, в мг',
                                               'Рибофлавин, в мг',
                                               'Ниацин, в мг', 'Аскорб кисл, в мг',
                                               'Холестерин, в мг',
                                               'НЖК, в г',
                                               'Ниационвый эквивалент, в мг',
                                               'Токоферол эквивалент, в мг'])

        # Считаем средний уровень микроэлементов
        list_of = ['Масса, гр','Углеводы, гр',
                    'Белки, гр', 'Жиры, гр',
                    'ККал', 'Микроэлементы', 'Вода, в г', 'МДС, в г',
                    'Крахмал, в г', 'Пищ вол, в г',
                    'Орган кислота, в г', 'Зола, в г',
                    'Натрий, в мг', 'Калий, в мг', 'Кальций, в мг',
                    'Магний, в мг', 'Фосфор, в мг', 'Железо, в мг',
                    'Ретинол, в мкг', 'Каротин, в мкг',
                    'Ретин экв, в мкг', 'Тиамин, в мг',
                    'Рибофлавин, в мг',
                    'Ниацин, в мг', 'Аскорб кисл, в мг',
                    'Холестерин, в мг',
                    'НЖК, в г',
                    'Ниационвый эквивалент, в мг',
                    'Токоферол эквивалент, в мг']

        mean2 = list()
        for i in list_of:
            exp = pd.to_numeric(food_weight[i])
            mean2.append(exp.mean())

        # Кривая попытка решить проблему с типом данных
        for name1 in list_of:
           for i in range(len(food_weight[name1])):
                food_weight[name1][i] = food_weight[name1][i].replace('.',',') + '\t'


        a = food_weight.groupby(['Дата',
                                 'Время',
                                 'Прием пищи']).agg({
                                  "Продукт": lambda tags: '\n'.join(tags),
                                  "Масса, гр": lambda tags: '\n'.join(tags),
                                  "Углеводы, гр": lambda tags: '\n'.join(tags),
                                  "Белки, гр": lambda tags: '\n'.join(tags),
                                  "Жиры, гр": lambda tags: '\n'.join(tags),
                                  "ККал": lambda tags: '\n'.join(tags),
                                  "Микроэлементы": lambda tags: '\n'.join(tags),
                                  "Вода, в г": lambda tags: '\n'.join(tags),
                                  "МДС, в г": lambda tags: '\n'.join(tags),
                                  "Крахмал, в г": lambda tags: '\n'.join(tags),
                                  "Пищ вол, в г": lambda tags: '\n'.join(tags),
                                  "Орган кислота, в г": lambda tags: '\n'.join(tags),
                                  "Зола, в г": lambda tags: '\n'.join(tags),
                                  "Натрий, в мг": lambda tags: '\n'.join(tags),
                                  "Калий, в мг": lambda tags: '\n'.join(tags),
                                  "Кальций, в мг": lambda tags: '\n'.join(tags),
                                  "Магний, в мг": lambda tags: '\n'.join(tags),
                                  "Фосфор, в мг": lambda tags: '\n'.join(tags),
                                  "Железо, в мг": lambda tags: '\n'.join(tags),
                                  "Ретинол, в мкг": lambda tags: '\n'.join(tags),
                                  "Каротин, в мкг": lambda tags: '\n'.join(tags),
                                  "Ретин экв, в мкг": lambda tags: '\n'.join(tags),
                                  "Тиамин, в мг": lambda tags: '\n'.join(tags),
                                  "Рибофлавин, в мг": lambda tags: '\n'.join(tags),
                                  "Ниацин, в мг": lambda tags: '\n'.join(tags),
                                  "Аскорб кисл, в мг": lambda tags: '\n'.join(tags),
                                  "Холестерин, в мг": lambda tags: '\n'.join(tags),
                                  "НЖК, в г": lambda tags: '\n'.join(tags),
                                  "Ниационвый эквивалент, в мг": lambda tags: '\n'.join(tags),
                                  "Токоферол эквивалент, в мг": lambda tags: '\n'.join(tags)}).reset_index()
        
        for i1 in range(len(a['Продукт'])):
            row = a['Продукт'][i1].split('\n')
            for i in range(len(row)):
                row[i] = f'{i+1}. ' + row[i]
            row = '\n'.join(row)
            a['Продукт'][i1] = row

        # Физическая активность
        activity1 = pd.DataFrame(L1, columns=['Дата', 'Время', 'Длительность, мин.',
                                              'Тип нагрузки','Пустое'])
        length1 = str(len(activity1['Время'])+3)
        activity2 = activity1.groupby(['Дата']).agg({
                                        'Время': lambda tags: '\n'.join(tags),
                                        'Длительность, мин.': lambda tags: '\n'.join(tags),
                                        'Тип нагрузки': lambda tags: '\n'.join(tags),
                                        'Пустое': lambda tags: '\n'.join(tags)})
        # Сон
        sleep1 = pd.DataFrame(L2, columns=['Дата', 'Время', 'Длительность, ч.'])

        sleep2 = sleep1.groupby(['Дата'
                                 ]).agg({'Время': lambda tags: '\n'.join(tags),
                                         'Длительность, ч.': lambda tags: '\n'.join(tags)})
        luck = pd.merge(left=activity2,
                        right=sleep2,
                        on="Дата", how='outer')

        luck["Дата1"] = pd.to_datetime(luck.index, format='%d.%m.%Y')                
        luck = luck.sort_values(by="Дата1")
        start1 = luck.index[0]
        end1 = luck.index[len(luck.index)-1]
        start1 = datetime.datetime.strptime(start1, '%d.%m.%Y')
        end1 = datetime.datetime.strptime(end1, '%d.%m.%Y')
        start1 = start1.strftime('%m/%d/%Y')
        end1 = end1.strftime('%m/%d/%Y')

        luck = luck.drop(["Дата1"], axis=1)

        ga = luck.index[0]
        gb = luck.index[len(luck.index)-1]
        ranges = pd.date_range(start=start1, end=end1)
        ranges1 = ranges.to_pydatetime()
        new_ranges = []
        for i in range(len(ranges1)):
            new_ranges.append(ranges1[i].strftime('%d.%m.%Y'))

        luck = luck.reindex(new_ranges)

        # Создаем общий Excel файл
        # можно добавить options={'strings_to_numbers': True} в writer
        # THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
        # my_file = os.path.join(THIS_FOLDER, '%s.xlsx' % session["username"])
        writer = pd.ExcelWriter('app\\%s.xlsx' % session["username"],
                                engine='xlsxwriter',
                                options={'strings_to_numbers': True,
                                         'default_date_format': 'dd/mm/yy'})
        a.to_excel(writer, sheet_name='Приемы пищи', startrow=0, startcol=0)

        luck.to_excel(writer, sheet_name='Физическая активность',
                           startrow=0, startcol=1)                                            
        writer.close()

        # Редактируем оформление приемов пищи
        wb = openpyxl.load_workbook('app\\%s.xlsx' % session["username"])
        sheet = wb['Приемы пищи']
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrapText=True)
                cell.alignment = cell.alignment.copy(vertical='center')


        for b in ['F','G','H','I','J','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH']:
            for i in range(2,((len(a['Микроэлементы'])+2))):
                k=i
                cs = sheet['%s' % b+str(k) ]
                cs.alignment = cs.alignment.copy(horizontal='left')

        for c in ['B','C','D']:
            for i in range(2,((len(a['Микроэлементы'])+2))):        
                k=i
                cs = sheet['%s' % c+str(k) ]
                cs.alignment = cs.alignment.copy(horizontal='center')


        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 7
        sheet.column_dimensions['D'].width = 13
        sheet.column_dimensions['E'].width = 50
        sheet.column_dimensions['F'].width = 13
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 20
        sheet.column_dimensions['K'].width = 20
        sheet.column_dimensions['L'].width = 20
        sheet.column_dimensions['M'].width = 20
        sheet.column_dimensions['N'].width = 20
        sheet.column_dimensions['O'].width = 20
        sheet.column_dimensions['P'].width = 20
        sheet.column_dimensions['R'].width = 20
        sheet.column_dimensions['S'].width = 20
        sheet.column_dimensions['T'].width = 20
        sheet.column_dimensions['O'].width = 20
        sheet.column_dimensions['U'].width = 20
        sheet.column_dimensions['V'].width = 20
        sheet.column_dimensions['W'].width = 20
        sheet.column_dimensions['X'].width = 20
        sheet.column_dimensions['Y'].width = 20
        sheet.column_dimensions['Z'].width = 20
        sheet.column_dimensions['Q'].width = 20
        sheet.column_dimensions['AA'].width = 20
        sheet.column_dimensions['AB'].width = 20
        sheet.column_dimensions['AC'].width = 20
        sheet.column_dimensions['AD'].width = 20
        sheet.column_dimensions['AE'].width = 20
        sheet.column_dimensions['AF'].width = 20
        sheet.column_dimensions['AG'].width = 30
        sheet.column_dimensions['AH'].width = 30
        sheet.column_dimensions['AI'].width = 20
        sheet.column_dimensions['AJ'].width = 20


        b1 = ws['B1']
        b1.fill = PatternFill("solid", fgColor="fafad2")
        c1 = ws['C1']
        c1.fill = PatternFill("solid", fgColor="fafad2")
        d1 = ws['D1']
        d1.fill = PatternFill("solid", fgColor="fafad2")
        e1 = ws['E1']
        e1.fill = PatternFill("solid", fgColor="fafad2")
        f1 = ws['F1']
        f1.fill = PatternFill("solid", fgColor="fafad2")
        g1 = ws['G1']
        g1.fill = PatternFill("solid", fgColor="fafad2")
        h1 = ws['H1']
        h1.fill = PatternFill("solid", fgColor="fafad2")
        i1 = ws['I1']
        i1.fill = PatternFill("solid", fgColor="fafad2")
        j1 = ws['J1']
        j1.fill = PatternFill("solid", fgColor="fafad2")

        m1 = ws['M1']
        m1.fill = PatternFill("solid", fgColor="fafad2")
        n1 = ws['N1']
        n1.fill = PatternFill("solid", fgColor="fafad2")
        o1 = ws['O1']
        o1.fill = PatternFill("solid", fgColor="fafad2")
        p1 = ws['P1']
        p1.fill = PatternFill("solid", fgColor="fafad2")
        q1 = ws['Q1']
        q1.fill = PatternFill("solid", fgColor="fafad2")
        r1 = ws['R1']
        r1.fill = PatternFill("solid", fgColor="fafad2")
        s1 = ws['S1']
        s1.fill = PatternFill("solid", fgColor="fafad2")
        t1 = ws['T1']
        t1.fill = PatternFill("solid", fgColor="fafad2")
        u1 = ws['U1']
        u1.fill = PatternFill("solid", fgColor="fafad2")
        v1 = ws['V1']
        v1.fill = PatternFill("solid", fgColor="fafad2")
        w1 = ws['W1']
        w1.fill = PatternFill("solid", fgColor="fafad2")
        x1 = ws['X1']
        x1.fill = PatternFill("solid", fgColor="fafad2")
        y1 = ws['Y1']
        y1.fill = PatternFill("solid", fgColor="fafad2")
        z1 = ws['Z1']
        z1.fill = PatternFill("solid", fgColor="fafad2")
        aa1 = ws['AA1']
        aa1.fill = PatternFill("solid", fgColor="fafad2")
        ab1 = ws['AB1']
        ab1.fill = PatternFill("solid", fgColor="fafad2")
        ac1 = ws['AC1']
        ac1.fill = PatternFill("solid", fgColor="fafad2")
        ad1 = ws['AD1']
        ad1.fill = PatternFill("solid", fgColor="fafad2")
        ae1 = ws['AE1']
        ae1.fill = PatternFill("solid", fgColor="fafad2")
        af1 = ws['AF1']
        af1.fill = PatternFill("solid", fgColor="fafad2")
        ah1 = ws['AH1']
        ah1.fill = PatternFill("solid", fgColor="fafad2")
        ag1 = ws['AG1']
        ag1.fill = PatternFill("solid", fgColor="fafad2")
        ws['AH1'].fill = PatternFill("solid", fgColor="fafad2")
        ws['L1'].fill = PatternFill("solid", fgColor="fafad2")

        i = 1
        for num in range(1,len(a['Микроэлементы'])+1):                
            if ws[f'B{num+1}'].value != ws[f'B{num}'].value:
                if i % 2 == 0:
                    ws[f'B{num+1}'].fill = PatternFill("solid", fgColor="f0f8ff")
                    i = i + 1
                else:
                    ws[f'B{num+1}'].fill = PatternFill("solid", fgColor="f0fff0")
                    i = i + 1                     
            else:
                ws[f'B{num+1}']._style = ws[f'B{num}']._style

        for i in ["C","D","E","F","G","H","I","J",'L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH']:
            for num in range(1,len(a['Микроэлементы'])+2):
                cell = ws[f'B{num}']
                ws[f'{i}{num}'].fill = PatternFill("solid", fgColor=cell.fill.start_color.index)      

        thin_border = Border(left=Side(style='hair'),
                             right=Side(style='hair'),
                             top=Side(style='hair'),
                             bottom=Side(style='hair'))

        no_border = Border(left=Side(border_style=None),
                             right=Side(border_style=None),
                             top=Side(border_style=None),
                             bottom=Side(border_style=None))                      

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        merged_cells_range = ws.merged_cells.ranges

        for merged_cell in merged_cells_range:
            merged_cell.shift(0, 2)
        ws.insert_rows(1, 2)

        # Разделяем основные показатели и микроэлементы
        ws['K3'].value = ''
        for i in range(len(a['Микроэлементы'])+3):
            i1 = str(i+1)
            ws[f'K{i1}'].border = no_border
        # Убираем форматирование первого столбца A1 и последнего AI
        for i in range(len(a['Микроэлементы'])+3):
            i1 = str(i+1)
            ws[f'A{i1}'].border = no_border
            ws[f'A{i1}'].value = ''
        for i in range(len(a['Микроэлементы'])+3):
            i1 = str(i+1)
            ws[f'AI{i1}'].border = no_border                     

        # Оформляем верхушки
        ws['A2'] = 'Приемы пищи'
        ws['A1'] = '%s' % fio[0][0]
        sheet.merge_cells('A1:AH1')
        ws['A2'].border = thin_border
        ws['A2'].fill = PatternFill("solid", fgColor="fafad2")
        ws['A2'].font = Font(bold=True)
        sheet.merge_cells('A2:AH2')
    

        length2 = str(len(a['Микроэлементы'])+5)
        length3 = str(len(a['Микроэлементы'])+6)
        sheet.merge_cells('C%s:E%s' % (length3,length3))
        ws['A%s' % length2] = 'Срденее по дням'
        ws['A%s' % length2].font = Font(bold=True)
        ws['B%s' % length3] = 'Дата'
        ws['B%s' % length3].font = Font(bold=True)
        ws['A%s' % length2].border = thin_border
        ws['A%s' % length2].fill = PatternFill("solid", fgColor="fafad2")
        ws['B%s' % length3].border = thin_border
        ws['B%s' % length3].fill = PatternFill("solid", fgColor="fafad2")
        ws['C%s' % length3].border = thin_border
        ws['C%s' % length3].fill = PatternFill("solid", fgColor="fafad2")
        
        # Проставляем внизу для средних по дням те же наименования, что и сверху
        mean21=['Масса, гр','Углеводы, гр',
                    'Белки, гр', 'Жиры, гр',
                    'ККал', '', 'Вода, в г', 'МДС, в г',
                    'Крахмал, в г', 'Пищ вол, в г',
                    'Орган кислота, в г', 'Зола, в г',
                    'Натрий, в мг', 'Калий, в мг', 'Кальций, в мг',
                    'Магний, в мг', 'Фосфор, в мг', 'Железо, в мг',
                    'Ретинол, в мкг', 'Каротин, в мкг',
                    'Ретин экв, в мкг', 'Тиамин, в мг',
                    'Рибофлавин, в мг',
                    'Ниацин, в мг', 'Аскорб кисл, в мг',
                    'Холестерин, в мг',
                    'НЖК, в г',
                    'Ниационвый эквивалент, в мг',
                    'Токоферол эквивалент, в мг']
        i = 0
        for c in ['F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH']:
            ws[f'{c}%s' % length3] = mean21[i]
            ws[f'{c}%s' % length3].border = thin_border
            ws[f'{c}%s' % length3].fill = PatternFill("solid", fgColor="fafad2")
            ws[f'{c}%s' % length3].font = Font(bold=True)
            i = i+1

        # Убираем закрашенные клетки пустого столбца K    
        length5 = str(len(a['Микроэлементы'])+8+len(date))
        ws['K%s' % length3]._style = copy(ws['K%s' % length5]._style)  
        ws['K%s' % length3].border = no_border

        # Выводим скользящее среднее
        path = os.path.dirname(os.path.abspath(__file__))
        db = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db)
        cur = con.cursor()
        i=7
        for d in date:
            sheet['B%s' % str(len(a['Микроэлементы'])+i)] = d[0]                 
            cur.execute('''SELECT avg(libra), avg(carbo), avg(prot), avg(fat), avg(energy), avg(water), avg(mds), avg(kr),
                           avg(pv), avg(ok), avg(zola), avg(na), avg(k), avg(ca), avg(mg), avg(p), avg(fe),
                           avg(a), avg(kar), avg(re), avg(b1), avg(b2), avg(rr), avg(ca), avg(hol), avg(nzhk), avg(ne),
                           avg(te) FROM favourites
                           WHERE user_id = ?
                           AND date = ? ''', (session['user_id'], d[0]))
            avg = cur.fetchall()
            sheet['F%s' % str(len(a['Микроэлементы'])+i)] = avg[0][0]
            sheet['G%s' % str(len(a['Микроэлементы'])+i)] = avg[0][1]
            sheet['H%s' % str(len(a['Микроэлементы'])+i)] = avg[0][2]
            sheet['I%s' % str(len(a['Микроэлементы'])+i)] = avg[0][3]
            sheet['J%s' % str(len(a['Микроэлементы'])+i)] = avg[0][4]
            sheet['L%s' % str(len(a['Микроэлементы'])+i)] = avg[0][5]
            sheet['M%s' % str(len(a['Микроэлементы'])+i)] = avg[0][6]
            sheet['N%s' % str(len(a['Микроэлементы'])+i)] = avg[0][7]
            sheet['O%s' % str(len(a['Микроэлементы'])+i)] = avg[0][8]
            sheet['P%s' % str(len(a['Микроэлементы'])+i)] = avg[0][9]
            sheet['Q%s' % str(len(a['Микроэлементы'])+i)] = avg[0][10]
            sheet['R%s' % str(len(a['Микроэлементы'])+i)] = avg[0][11]
            sheet['S%s' % str(len(a['Микроэлементы'])+i)] = avg[0][12]
            sheet['T%s' % str(len(a['Микроэлементы'])+i)] = avg[0][13]
            sheet['U%s' % str(len(a['Микроэлементы'])+i)] = avg[0][14]
            sheet['V%s' % str(len(a['Микроэлементы'])+i)] = avg[0][15]
            sheet['W%s' % str(len(a['Микроэлементы'])+i)] = avg[0][16]
            sheet['X%s' % str(len(a['Микроэлементы'])+i)] = avg[0][17]
            sheet['Y%s' % str(len(a['Микроэлементы'])+i)] = avg[0][18]
            sheet['Z%s' % str(len(a['Микроэлементы'])+i)] = avg[0][19]
            sheet['AA%s' % str(len(a['Микроэлементы'])+i)] = avg[0][20]
            sheet['AB%s' % str(len(a['Микроэлементы'])+i)] = avg[0][21]
            sheet['AC%s' % str(len(a['Микроэлементы'])+i)] = avg[0][22]
            sheet['AD%s' % str(len(a['Микроэлементы'])+i)] = avg[0][23]
            sheet['AE%s' % str(len(a['Микроэлементы'])+i)] = avg[0][24]
            sheet['AF%s' % str(len(a['Микроэлементы'])+i)] = avg[0][25]
            sheet['AG%s' % str(len(a['Микроэлементы'])+i)] = avg[0][26]
            sheet['AH%s' % str(len(a['Микроэлементы'])+i)] = avg[0][27]
            i = i + 1     
        con.close()

        # Выравниваем скользящее среднее по левому краю
        length31 = len(a['Микроэлементы'])+7
        length4 = len(a['Микроэлементы'])+7+len(date)

        for a in ['F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH']:
            for i in range(length31,length4):
                sheet[f'{a}{i}'].alignment = sheet[f'{a}{i}'].alignment.copy(horizontal = 'left')
        for a in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH']:
            sheet[f'{a}3'].alignment = sheet[f'{a}3'].alignment.copy(horizontal = 'left')    

        ws.protection.set_password('test')
        wb.save('app\\%s.xlsx' % session["username"])
        wb.close()

        # Форматируем физическую активность как надо
        wb = openpyxl.load_workbook('app\\%s.xlsx' % session["username"])
        sheet1 = wb['Физическая активность']

        for row in sheet1.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrapText=True)
                cell.alignment = cell.alignment.copy(vertical='center')
                cell.alignment = cell.alignment.copy(horizontal='left')

        thin_border = Border(left=Side(style='hair'),
                             right=Side(style='hair'),
                             top=Side(style='hair'),
                             bottom=Side(style='hair'))

        no_border = Border(left=Side(border_style=None),
                             right=Side(border_style=None),
                             top=Side(border_style=None),
                             bottom=Side(border_style=None))

        for row in sheet1.iter_rows():
            for cell in row:
                cell.border = thin_border 

        merged_cells_range = sheet1.merged_cells.ranges

        for merged_cell in merged_cells_range:
            merged_cell.shift(0, 2)
        sheet1.insert_rows(1, 2)

        sheet1['A1'] = '%s' % fio[0][0]
        sheet1['A2'] = 'Физическая активность'
        sheet1['G2'] = 'Сон'
        sheet1.merge_cells('A1:H1')

        sheet1.column_dimensions['A'].width = 25
        sheet1.column_dimensions['B'].width = 13
        sheet1.column_dimensions['C'].width = 13
        sheet1.column_dimensions['D'].width = 20
        sheet1.column_dimensions['E'].width = 25
        sheet1.column_dimensions['F'].width = 13
        sheet1.column_dimensions['G'].width = 13
        sheet1.column_dimensions['H'].width = 20
               
        b1 = sheet1['B3']
        b1.fill = PatternFill("solid", fgColor="fafad2")
        c1 = sheet1['C3']
        c1.fill = PatternFill("solid", fgColor="fafad2")
        d1 = sheet1['D3']
        d1.fill = PatternFill("solid", fgColor="fafad2")
        e1 = sheet1['E3']
        e1.fill = PatternFill("solid", fgColor="fafad2")

        g1 = sheet1['G3']
        g1.fill = PatternFill("solid", fgColor="fafad2")
        sheet1['H3'].fill = PatternFill("solid", fgColor="fafad2")

        # Разделяем физическую нагрузку и сон, также убираем форматирование с первого столбца A1
        # убираем мелкие дефекты
        sheet1['F3'].value = ''
        sheet1['C3'].value = 'Время'
        sheet1['G3'].value = 'Время'
        for i in range(3,len(luck['Длительность, ч.'])+4):
            i1 = str(i)
            sheet1[f'F{i1}'].border = no_border
            
        for i in range(3,len(luck['Длительность, ч.'])+4):
            i1 = str(i)
            sheet1[f'A{i1}'].border = no_border
        # Корректируем верхушки
        sheet1['A2'].fill = PatternFill("solid", fgColor="fafad2")
        sheet1['G2'].fill = PatternFill("solid", fgColor="fafad2")
        sheet1['A2'].border = thin_border
        sheet1['G2'].border = thin_border 

        sheet1['A2'].font = Font(bold=True)
        sheet1['G2'].font = Font(bold=True)
        for i in range(4, len(luck['Время_x'])+4):
            sheet1[f'B{i}'].font = Font(bold=False)

        # Закрашиваем строки через одну
        k = 1
        for abc in ['B','C','D','E','G','H']:
            for i in range(4,len(luck['Длительность, ч.'])+4):
                if k % 2 == 0:
                    sheet1[f'{abc}{i}'].fill = PatternFill('solid', fgColor='f0f8ff')
                    k = k + 1
                else:
                    sheet1[f'{abc}{i}'].fill = PatternFill('solid', fgColor='f0fff0')
                    k = k + 1

        # Устанавливаем пароль на лист и сохраняем
        sheet1.protection.set_password('test')
        wb.save('app\\%s.xlsx' % session["username"])
        wb.close()

        # Отправляем по почте
        msg = Message('ДиаКомпаньон', sender = 'teos.sicrets@gmail,com', recipients=mail1)
        msg.subject = "Никнейм пользователя: %s" % session["username"]
        msg.body = 'Электронный отчет'
        with app.open_resource('%s.xlsx' % session["username"]) as attach:
            msg.attach('%s.xlsx' % session["username"], 'sheet/xlsx',
                       attach.read())
        mail.send(msg)

    return redirect(url_for('lk'))


if __name__ == '__main__':
    app.run(debug=True)
